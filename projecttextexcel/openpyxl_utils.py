import warnings
from openpyxl.utils import range_boundaries, get_column_letter


def get_merged_openpyxl_cell(openpyxl_ws, openpyxl_cell):
    try:
        merged_range = [s for s in openpyxl_ws.merged_cells.ranges if openpyxl_cell.coordinate in s]
    except AttributeError:
        merged_range = []
        warnings.warn(f"Workbook's merged cells can't be parsed in read-only mode.")
    if merged_range:
        return openpyxl_ws.cell(merged_range[0].min_row, merged_range[0].min_col)
    else:
        openpyxl_cell


def get_cell_type(cell):
    """ Determine the type of a cell's value. """
    if cell.data_type == 'n':
        return "float" if '.' in str(cell.value) else "int"
    elif cell.data_type == 'd':
        return "datetime"
    elif cell.data_type == 's':
        return "str"
    elif cell.data_type == 'f':
        return 'formula'
    else:
        return 'unknown'


def range_generator(openpyxl_ws, table_range=None):
    """
    Generator function to iterate over a range in the worksheet.
    Yields (row_index, column_index, cell).
    """
    if table_range is not None:
        min_col, min_row, max_col, max_row = range_boundaries(table_range)
    else:
        min_col, min_row, max_col, max_row = 1, 1, None, None
    # Iterate through columns and rows
    for row_idx, row in enumerate(openpyxl_ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col), start=min_row):
        for col_idx, cell in enumerate(row, start=min_col):
            yield row_idx, col_idx, cell


def is_cell_empty(openpyxl_cell):
    return openpyxl_cell is None or openpyxl_cell.value is None or openpyxl_cell.value == ""


def get_excel_coordinate(min_col, min_row, max_col=None, max_row=None):
    if max_col is None and max_row is None:
        return f"{get_column_letter(min_col)}{min_row}"
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
