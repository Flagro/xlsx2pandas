import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries
from .utils import BaseTableDetector


def is_empty(cell):
    return cell is None or cell.value is None or cell.value == ""


class TableDetector(BaseTableDetector):
    def find_table_end(self, openpyxl_ws, start_row, start_col):
        ws_range = openpyxl_ws.calculate_dimension(force=True)
        _, ws_max_row, _, ws_max_col = range_boundaries(ws_range)
        max_row = start_row
        max_col = start_col
        empty_row_count = 0
        empty_col_count = 0

        # Iterate over rows
        for i in range(start_row, ws_max_row + 1):
            if all(is_empty(openpyxl_ws.cell(row=i, column=j)) for j in range(start_col, max_col + 1)):
                empty_row_count += 1
            else:
                empty_row_count = 0
                max_row = i

            if empty_row_count > 1:
                break

        # Iterate over columns
        for j in range(start_col, ws_max_col + 1):
            if all(is_empty(openpyxl_ws.cell(row=i, column=j)) for i in range(start_row, max_row + 1)):
                empty_col_count += 1
            else:
                empty_col_count = 0
                max_col = j

            if empty_col_count > 1:
                break

        return max_row, max_col

    def get_table_ranges(self, openpyxl_ws, **kwargs):
        tables = []
        visited = set()

        for row_idx, row in enumerate(openpyxl_ws.iter_rows(), start=1):
            for col_idx, cell in enumerate(row, start=1):
                cell_coord = f"{get_column_letter(col_idx)}{row_idx}"

                if cell_coord in visited or is_empty(cell):
                    continue

                end_row, end_col = self.find_table_end(openpyxl_ws, row_idx, col_idx)
                table_range = f"{get_column_letter(col_idx)}{row_idx}:{get_column_letter(end_col)}{end_row}"
                tables.append(table_range)

                # Mark cells as visited
                for r in range(row_idx, end_row + 1):
                    for c in range(col_idx, end_col + 1):
                        visited.add(f"{get_column_letter(c)}{r}")

        return tables
