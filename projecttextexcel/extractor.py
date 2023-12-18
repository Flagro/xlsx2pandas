import pandas as pd
from typing import Optional, List, Dict, Union
from openpyxl import load_workbook
from .table_detection.strategies.utils import get_strategy_parser
from .utils import get_in_memory_file, prettify_workbook_dataframes_output


def get_df(file_path, 
           sheet_names: Optional[Union[str, List[str]]], 
           table_detection_strategy='general',
           header_separator_strategy='general',
           prettify_output=True,
           **kwargs) -> Union[List[pd.DataFrame], Dict[str, List[pd.DataFrame]], Dict[str, pd.DataFrame], pd.DataFrame]:
    """
    Returns a list or a single pandas dataframe of the extracted data from the xlsx/xlsm file.
    """ 
    
    strategy_parser = get_strategy_parser(table_detection_strategy)

    in_memory_file = get_in_memory_file(file_path)
    wb = load_workbook(in_memory_file, read_only=True)

    if isinstance(sheet_names, str):
        sheets = [sheet_names]
    elif isinstance(sheet_names, list):
        sheets = sheet_names
    else:
        sheets = wb.sheetnames

    dataframes_dict = dict()

    for sheet in sheets:
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet {sheet} not found.")
        else:
            ws = wb[sheet]
            sheet_dataframes = strategy_parser.parse(ws, **kwargs)
            dataframes_dict[sheet] = sheet_dataframes
    wb.close()

    if prettify_output:
        dataframes_dict = prettify_workbook_dataframes_output(dataframes_dict)

    return dataframes_dict
