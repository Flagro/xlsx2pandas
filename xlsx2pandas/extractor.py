import warnings
import pandas as pd
from typing import Optional, List, Dict, Union
from openpyxl import load_workbook
from .table_detection.utils import get_table_detector
from .header_separation.utils import get_header_separator
from .dataframe_construction.utils import get_dataframe_constructor
from .utils import get_in_memory_file, prettify_workbook_dataframes_output


def get_df(file_path, 
           sheet_names: Optional[Union[str, List[str]]]=None, 
           table_detection_strategy='general',
           header_separator_strategy='general',
           dataframe_construction_strategy='general',
           prettify_output=True,
           **kwargs) -> Union[List[pd.DataFrame], Dict[str, List[pd.DataFrame]], Dict[str, pd.DataFrame], pd.DataFrame]:
    """
    Returns a list or a single pandas dataframe of the extracted data from the xlsx/xlsm file.
    """ 
    
    table_detector = get_table_detector(table_detection_strategy)
    header_separator = get_header_separator(header_separator_strategy)
    dataframe_constructor = get_dataframe_constructor(dataframe_construction_strategy)

    in_memory_file = get_in_memory_file(file_path)
    wb = load_workbook(in_memory_file, read_only=True)

    if isinstance(sheet_names, str):
        sheets = [sheet_names]
    elif isinstance(sheet_names, list):
        sheets = sheet_names
    else:
        sheets = wb.sheetnames

    dataframes_dict = dict()

    for sheet in sheets:
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet {sheet} not found.")
        else:
            ws = wb[sheet]
            # Get the table ranges in a list
            try:
                sheet_table_ranges = table_detector.get_table_ranges(ws, **kwargs)
            except Exception as e:
                sheet_table_ranges = []
                warnings.warn(f"Sheet {sheet} could not be parsed. Details: {e}")
            
            # Get header rows counters in a list
            sheet_header_rows_cnt = []
            for table_range in sheet_table_ranges:
                try:
                    header_rows_cnt = header_separator.get_header_rows_cnt(ws, table_range, **kwargs)
                except Exception as e:
                    header_rows_cnt = 1
                    warnings.warn(f"Sheet {sheet} could not be parsed for header rows. Defaulting to 1 header row. Details: {e}")
                sheet_header_rows_cnt.append(header_rows_cnt)

            # Construct the dataframes from the table ranges and header rows counters
            sheet_dataframes = []
            for table_range, header_rows_cnt in zip(sheet_table_ranges, sheet_header_rows_cnt):
                try:
                    df = dataframe_constructor.construct_dataframe(ws, table_range, header_rows_cnt, **kwargs)
                    sheet_dataframes.append(df)
                except Exception as e:
                    warnings.warn(f"Error constructing a DataFrame on sheet {sheet}, table range {table_range}, header rows cnt {header_rows_cnt}. Details: {e}")
            dataframes_dict[sheet] = sheet_dataframes
    wb.close()

    if prettify_output:
        dataframes_dict = prettify_workbook_dataframes_output(dataframes_dict)

    return dataframes_dict
